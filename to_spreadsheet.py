import openpyxl
from wiki_sp500_stock_list import ScrapeWikiSP500
from yahoo_stock_scraper import YahooStockQuote



sp500 = ScrapeWikiSP500()
sp500_holdings = sp500.sp_500_holdings

stock_symbols = []
for i in range(10):
    stock_symbols.append(sp500_holdings[i]['symbol'])


prices = []
for symbol in stock_symbols:
    quote = YahooStockQuote(symbol)
    o_price = quote.stock_info['open price']
    c_price = quote.stock_info['cur. price']
    prices.append({symbol: (o_price, c_price)})



wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Stock Symbol'
sheet['B1'] = 'open'
sheet['C1'] = 'close'

for i, stock in enumerate(prices):
    symbol, info = list(stock.items())[0]
    sheet.cell(column=1, row=i+2).value = symbol
    sheet.cell(column=2, row=i+2).value = info[0]
    sheet.cell(column=3, row=i+2).value = info[1]

wb.save("test_excel.xslx")
